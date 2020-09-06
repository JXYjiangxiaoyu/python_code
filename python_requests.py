#!/usr/bin/python3
#-*-编码：ascii-*-
'''
接口测试:第三方库--别人写好封装好，你可以直接拿来的用的功能=requests--参数传入用字典格式传入
两个步骤:
1、安装--pip自动下载安装第三方库=pip install requests --user
2、导入--Python文件里--范围内可以直接使用
'''
'''
import requests  # 导入第三方库
# 注册接口
qcd_url = "http://120.78.128.25:8766/futureloan/member/register"  # 接口地址
qcd_body = {"mobile_phone":"18385632662","pwd":"lemon666"}
qcd_headers = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
result = requests.post(url=qcd_url,json=qcd_body,headers = qcd_headers)
# print(result.json().get("data").get("id"))   # 获取响应消息的id
'''

# 获取token
# 方式1：
# result.json().get("data").get("token_info").get("token")
# 获取id
# result.json().get("data").get("id")

'''
注意：知识点：sonpath.jsonpath(reps.json(),"$..id")[0]
解释：..id 表示不管前面的有多少节点的参数，要找到id才可以。
'''
'''
# 方式二。使用json提取器：jsonpath , $..token：表示获取token,两个点表示匹配任意节点。
import jsonpath  # 导入第三方库：jsonpath
token= jsonpath.jsonpath(result.json(),"$..token")
id = jsonpath.jsonpath(result.json(),"$..id")
'''

# 使用内置函数pprint
# import  pprint  # 导入pprint
# pprint.pprint(result.json())


'''
# 第三方库：openpyxl   ----- 读取、回写
1.安装：pip isntall openpyxl
2.导入：import  openpyxl
# excel表格的常用操作：三大对象
1、工作簿对象
2、表单---sheet
3、单元格---cell
'''

'''
=====================================
定义函数的三大步骤：
1.实现功能
2.参数化的值
3.返回值----别人需要从你这里得到的值
=====================================
'''

'''
发送接口请求
'''
import requests
import openpyxl  # 导入第三方库


# 1.函数一，请求接口
def api_request(api_url, api_data):
    qcd_header = {'X-Lemonban-Media-Type': 'lemonban.v2','Content-Type': 'application/json'}
    response = requests.post(url=api_url, json=api_data, headers=qcd_header)  # 返回值，响应体
    return response.json()  # 返回响应结果


# 2.函数2：对excel表格进行读取
def read_data(fileName, sheetName):
    wb = openpyxl.load_workbook(fileName)  # 加载工作簿对象
    sheet = wb[sheetName]  # 获取到表单
    case_list = []  # 定义一个空列表，用来接收用例
    max_row = sheet.max_row  # 获取表单最大行数

    for i in range(2, max_row + 1):
        case = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,  # 行、列、找到对应的单元格
            data=sheet.cell(row=i, column=6).value,  # 参数
            expected=sheet.cell(row=i, column=7).value
        )  # 一个字典是一个用例

        case_list.append(case)
    return case_list  # 返回值


# 2.1调用遍历用例方法
# cases = read_data("test_case_api.xlsx","register")
# print(f'========{cases}===========')


# 3.回写测试结果到表格
def write_result(fileName, sheetName, row, column, final_result):
    wb = openpyxl.load_workbook(fileName)  # 加载工作簿对象
    sheet = wb[sheetName]  # 获取到表单
    sheet.cell(row=row, column=column).value = final_result
    wb.save(fileName)


# 4.读取结果
def execute_func(fileName, sheetName):
    cases = read_data(fileName, sheetName)  # 变量接收函数的值

    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")  # 字典的取值方式1
        data = case["data"]  # 字典的取值方式2
        data = eval(data)  # 去掉引号的函数

        expected = case.get("expected")
        expected2 = eval(expected)
        real_result = api_request(api_url=url, api_data=data)  # 执行结果
        real_msg = real_result.get("msg")
        expected_msg = expected2.get("msg")
        print("执行的结果是：{}".format(real_msg))
        print("预期的结果是：{}".format(expected_msg))


        if real_msg == expected_msg:
            print(f'第{case_id}条用例执行通过')
        else:
            print(f'第{case_id}条用例执行不通过')

# 4.1 调用execute_func()
execute_func("test_case_api.xlsx","register")



# =========================1.request 面试题=================================
# 1、乱码
# #2、页面内容不对
'''
import requests
baidu_url= "https://www.baidu.com/"
ningmengban_url ="https://www.baidu.com/s"
baidu_headers = {"User-Agent":
	"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:80.0) Gecko/20100101 Firefox/80.0"}
ningmenbing_par = {"wd":"柠檬班"}


res = requests.get(url=ningmengban_url,headers = baidu_headers,params=ningmenbing_par)
print(res.text) #文本格式结果自动进行解码--大部分的页-808
print(res.content.decode ( "utf8")) #手动指定解码
'''
