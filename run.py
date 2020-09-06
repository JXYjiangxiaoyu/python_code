#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
发送接口请求
'''
import requests
import openpyxl  # 导入第三方库


# 1.函数一，请求接口
def api_request(api_url, api_data):
    qcd_header = {'X-Lemonban-Media-Type': 'lemonban.v2','Content-Type': 'application/json'}
    response = requests.post(url=api_url, json=api_data, headers=qcd_header)  # 返回值，响应体
    return response.json()  # 返回响应结果


# 2.函数2：对excel表格进行读取
def read_data(fileName, sheetName):
    wb = openpyxl.load_workbook(fileName)  # 加载工作簿对象
    sheet = wb[sheetName]  # 获取到表单
    case_list = []  # 定义一个空列表，用来接收用例
    max_row = sheet.max_row  # 获取表单最大行数

    for i in range(2, max_row + 1):
        case = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,  # 行、列、找到对应的单元格
            data=sheet.cell(row=i, column=6).value,  # 参数
            expected=sheet.cell(row=i, column=7).value
        )  # 一个字典是一个用例

        case_list.append(case)
    return case_list  # 返回值


# 2.1调用遍历用例方法
# cases = read_data("test_case_api.xlsx","register")
# print(f'========{cases}===========')


# 3.回写测试结果到表格
def write_result(fileName, sheetName, row, column, final_result):
    wb = openpyxl.load_workbook(fileName)  # 加载工作簿对象
    sheet = wb[sheetName]  # 获取到表单
    sheet.cell(row=row, column=column).value = final_result
    wb.save(fileName)


# 4.读取结果
def execute_func(fileName, sheetName):
    cases = read_data(fileName, sheetName)  # 变量接收函数的值

    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")  # 字典的取值方式1
        data = case["data"]  # 字典的取值方式2
        data = eval(data)  # 去掉引号的函数

        expected = case.get("expected")
        expected2 = eval(expected)
        real_result = api_request(api_url=url, api_data=data)  # 执行结果
        real_msg = real_result.get("msg")
        expected_msg = expected2.get("msg")
        print("执行的结果是：{}".format(real_msg))
        print("预期的结果是：{}".format(expected_msg))


        # if real_msg == expected_msg:
        #     print(f'第{case_id}条用例执行通过')
        # else:
        #     print(f'第{case_id}条用例执行不通过')

# 4.1 调用execute_func()
execute_func("test_case_api.xlsx","register")
